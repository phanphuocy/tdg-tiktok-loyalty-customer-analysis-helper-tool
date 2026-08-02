from openpyxl import load_workbook
import sqlite3
from pathlib import Path
import pandas as pd

from labelling import label_tuples_from_csv, add_label_header_tuple

from queries import *

# orders_filename = 'Tất cả đơn hàng-2026-06-01-07_52.xlsx'
orders_filename = 'Tất cả đơn hàng-2026-08-01-13_45.xlsx'
# output_filename = 'data.db'
output_filename = 'data_th8.db'
orders_active_sheet = 'OrderSKUList'

# product_labels_filename = 'Product_Additional_Labels - Product_AddtionalLabels_Labels.csv'
# product_labels_filename = 'Product_Additional_Labels2 - Product_AddtionalLabels_Labels2.csv'
# product_labels_filename = 'Product_Additional_Labels2 - Product_AddtionalLabels_Labels3.csv'
product_labels_filename = 'Product_Additional_Labels3 - roduct_AddtionalLabels_Labels4.csv'
# product_labels_sheet = 'Product_AddtionalLabels_Labels'
# product_labels_sheet = 'Product_AddtionalLabels_Labels2'
# product_labels_sheet = 'Product_AddtionalLabels_Labels3'
product_labels_sheet = 'Product_AddtionalLabels_Labels4'

orders_workbook = load_workbook(filename=orders_filename)
orders_sheet = orders_workbook[orders_active_sheet]

# 2. Extract rows (Separating headers from data)
orders_rows = list(orders_sheet.iter_rows(values_only=True))
if not orders_rows:
    print("The Excel sheet is empty.")
    exit()

orders_headers = orders_rows[0]  # First row contains column names
orders_data_rows = orders_rows[2:]  # Remaining rows contain the actual data

labelled_tuples, missing_ids = label_tuples_from_csv(
    csv_path=product_labels_filename,
    tuples=orders_data_rows,
    id_index=5,
)

labelled_headers = add_label_header_tuple(orders_headers)

print(f"Missing ids: {len(missing_ids)}")

# 3. Clean headers for SQL safety (replace spaces/special chars with underscores)
clean_headers = [
    str(h).strip().replace(" ", "_").replace("-", "_") if h else f"column_{i}"
    for i, h in enumerate(labelled_headers)
]

# 4. Connect to SQLite database (creates 'data.db' file if it doesn't exist)
file_path = Path(output_filename)

if file_path.is_file():
    file_path.unlink()

conn = sqlite3.connect(output_filename)
cursor = conn.cursor()

# 5. Dynamically create the SQL table based on Excel headers
# All columns are set to TEXT for safety, but SQLite handles dynamic typing
columns_query = ", ".join([f"[{name}] TEXT" for name in clean_headers])
create_table_query = f"CREATE TABLE IF NOT EXISTS excel_data ({columns_query})"
cursor.execute(create_table_query)

# 6. Build the INSERT query and inject the data rows
placeholders = ", ".join(["?"] * len(clean_headers))
insert_query = f"INSERT INTO excel_data VALUES ({placeholders})"

cursor.executemany(insert_query, labelled_tuples)
print(f"Successfully converted {len(labelled_tuples)} rows into '{output_filename}'!")

cursor.execute(CLEAR_OUT_GIFT_DATA)

### ------------------------------------------ ###
### ----------- TOTAL ORDERS TABLE ----------- ###
### ------------------------------------------ ###
cursor.execute(PRE_CREATE_TOTAL_ORDERS_TABLE)
cursor.execute(CREATE_TOTAL_ORDERS_TABLE)

cursor.execute("SELECT * FROM total_orders_data")
total_orders_data = cursor.fetchall()
print(f"Successfully added {len(total_orders_data)} rows into 'total_orders_data' table!")


### --------------------------------------------- ###
### ----------- TOTAL CUSTOMERS TABLE ----------- ###
### --------------------------------------------- ###
cursor.execute(PRE_CREATE_TOTAL_CUSTOMERS_TABLE)
cursor.execute(CREATE_TOTAL_CUSTOMERS_TABLE)
cursor.execute(PRE_UPDATE_TOTAL_CUSTOMERS_TABLE_WITH_FUNNEL_COL)
cursor.execute(UPDATE_TOTAL_CUSTOMERS_TABLE_WITH_FUNNEL_COL)
cursor.execute(PRE_UPDATE_TOTAL_CUSTOMERS_TABLE_WITH_SWITCHINGTIME_COL)
cursor.execute(UPDATE_TOTAL_CUSTOMERS_TABLE_WITH_SWITCHINGTIME_COL)
cursor.execute(PRE_UPDATE_TOTAL_CUSTOMERS_TABLE_WITH_DAYTOSWITCH_COL)
cursor.execute(UPDATE_TOTAL_CUSTOMERS_TABLE_WITH_DAYTOSWITCH_COL)
cursor.execute(CREATE_LOYAL_SWITCHING_CUSTOMERS_FILTER_TABLE)

cursor.execute("SELECT * FROM total_customers_data")
total_customers_data = cursor.fetchall()
print(f"Successfully added {len(total_customers_data)} rows into 'total_customers_data' table!")


### ----------------------------------------------------- ###
### ----------- TOTAL CUSTOMERS LOYALTY TABLE ----------- ###
### ----------------------------------------------------- ###
cursor.execute(PRE_CREATE_PIVOT_LOYALTY_TABLE)
cursor.execute(CREATE_PIVOT_LOYALTY_TABLE)

cursor.execute("SELECT * FROM pivot_loyalty_tier")
pivot_loyalty_tier = cursor.fetchall()
print(f"Successfully added {len(pivot_loyalty_tier)} rows into 'pivot_loyalty_tier' table!")


### ----------------------------------------------- ###
### ----------- MONTHLY CUSTOMERS TABLE ----------- ###
### ----------------------------------------------- ###
cursor.execute(PRE_CREATE_MONTHLY_CUSTOMERS_TABLE)
cursor.execute(CREATE_MONTHLY_CUSTOMERS_TABLE)

cursor.execute("SELECT * FROM monthly_customers_data")
monthly_customers_data = cursor.fetchall()
print(f"Successfully added {len(monthly_customers_data)} rows into 'monthly_customers_data' table!")


### ------------------------------------------------------- ###
### ----------- MONTHLY CUSTOMERS LOYALTY TABLE ----------- ###
### ------------------------------------------------------- ###

cursor.execute(PRE_CREATE_PIVOT_MONTHLY_CUSTOMERS_ACQUISITION)
cursor.execute(CREATE_PIVOT_MONTHLY_CUSTOMERS_ACQUISITION)

cursor.execute("SELECT * FROM pivot_monthly_customers_acquisition")
pivot_monthly_customers_acquisition = cursor.fetchall()
print(f"Successfully added {len(pivot_monthly_customers_acquisition)} rows into 'pivot_monthly_customers_acquisition' table!")

### --------------------------------------------- ###
### ----------- SWITCHING TABLE ----------- ###
### --------------------------------------------- ###
cursor.execute(CREATE_SWITCH_TOTAL_SWITCHING_STATUS)
cursor.execute(CREATE_SWITCH_MONTHLY_SWITCHING_STATUS_LOYAL)
cursor.execute(CREATE_TOTAL_FUNNEL_GROUP_PIVOT_TABLE)
cursor.execute(CREATE_SWITCH_MONTHLY_FUNNEL_GROUP)
cursor.execute(CREATE_SWITCH_MONTHLY_FUNNEL_GROUP_SWITCHER)
print("Ran all switching table generators")

### ------------------------------------------- ###
### ----------- OTHERS PIVOT TABLES ----------- ###
### ------------------------------------------- ###

cursor.execute(CREATE_PIVOT_MONTHLY_PRODUCTS)

cursor.execute("SELECT * FROM pivot_monthly_products")
pivot_monthly_products = cursor.fetchall()
print(f"Successfully added {len(pivot_monthly_products)} rows into 'pivot_monthly_products' table!")


### --------------------------------------------- ###
### ------------ RETENTION RATE TABLE ----------- ###
### --------------------------------------------- ###

cursor.execute("SELECT * FROM pivot_monthly_customers_acquisition")
pivot_monthly_customers_acquisition = cursor.fetchall()

columns = [desc[0] for desc in cursor.description]
df = pd.DataFrame(pivot_monthly_customers_acquisition, columns=columns)
df = df.sort_values('Order_Month').reset_index(drop=True)
 
new = df[df['Acquisition_Type']=='Newly Accquired'][['Order_Month','Num_Of_Customers']].set_index('Order_Month')
ret = df[df['Acquisition_Type']=='Return From Previous Month'][['Order_Month','Num_Of_Customers']].set_index('Order_Month')
 
months = sorted(df['Order_Month'].unique())
records = []
total_new = 0
 
for m in months:
    n = new.loc[m, 'Num_Of_Customers'] if m in new.index else 0
    r = ret.loc[m, 'Num_Of_Customers'] if m in ret.index else 0
    pool = total_new
    rate = round(r / pool * 100, 2) if pool > 0 else None
    records.append({
        'Order_Month': m,
        'New_Customers': n,
        'Legacy_Pool': pool,
        'Returning_Customers': r,
        'Retention_Rate_Pct': rate
    })
    total_new += n
 
result_df = pd.DataFrame(records)
result_df.to_sql('monthly_retention_rate', conn, if_exists='replace', index=False)
print(f"Successfully added {len(records)} rows into 'monthly_retention_rate' table!")


# 7. Commit changes and close the connection
conn.commit()
conn.close()